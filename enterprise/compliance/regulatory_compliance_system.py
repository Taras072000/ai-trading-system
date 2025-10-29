"""
Enterprise Regulatory Compliance System - Система регуляторного соответствия
Обеспечивает соблюдение требований регуляторов и автоматическую отчетность
"""

import asyncio
import json
import time
import hashlib
from typing import Dict, List, Optional, Any, Tuple, Union
from dataclasses import dataclass, asdict
from datetime import datetime, timedelta
from enum import Enum
import logging
import redis.asyncio as redis
from collections import defaultdict, deque
import uuid

# Криптография и безопасность
import cryptography
from cryptography.fernet import Fernet
from cryptography.hazmat.primitives import hashes
from cryptography.hazmat.primitives.kdf.pbkdf2 import PBKDF2HMAC
import base64

# Отчеты и документы
import pandas as pd
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import inch
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

# Email и уведомления
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders

# Метрики и мониторинг
from prometheus_client import Counter, Histogram, Gauge

# Математические библиотеки
import numpy as np
from scipy import stats

class RegulatorType(Enum):
    """Типы регуляторов"""
    SEC = "sec"  # Securities and Exchange Commission
    CFTC = "cftc"  # Commodity Futures Trading Commission
    FINRA = "finra"  # Financial Industry Regulatory Authority
    FCA = "fca"  # Financial Conduct Authority (UK)
    ESMA = "esma"  # European Securities and Markets Authority
    ASIC = "asic"  # Australian Securities and Investments Commission
    FSA = "fsa"  # Financial Services Agency (Japan)
    CBRF = "cbrf"  # Central Bank of Russia

class ComplianceRuleType(Enum):
    """Типы правил соответствия"""
    POSITION_LIMIT = "position_limit"
    RISK_LIMIT = "risk_limit"
    REPORTING = "reporting"
    KYC = "kyc"  # Know Your Customer
    AML = "aml"  # Anti-Money Laundering
    MARKET_ABUSE = "market_abuse"
    BEST_EXECUTION = "best_execution"
    RECORD_KEEPING = "record_keeping"
    CAPITAL_ADEQUACY = "capital_adequacy"

class ViolationSeverity(Enum):
    """Серьезность нарушения"""
    LOW = "low"
    MEDIUM = "medium"
    HIGH = "high"
    CRITICAL = "critical"

class ReportType(Enum):
    """Типы отчетов"""
    DAILY_TRADING = "daily_trading"
    WEEKLY_RISK = "weekly_risk"
    MONTHLY_COMPLIANCE = "monthly_compliance"
    QUARTERLY_FINANCIAL = "quarterly_financial"
    ANNUAL_AUDIT = "annual_audit"
    INCIDENT_REPORT = "incident_report"
    REGULATORY_FILING = "regulatory_filing"

@dataclass
class ComplianceRule:
    """Правило соответствия"""
    id: str
    name: str
    type: ComplianceRuleType
    regulator: RegulatorType
    description: str
    parameters: Dict[str, Any]
    is_active: bool = True
    created_at: datetime = None
    
    def __post_init__(self):
        if self.created_at is None:
            self.created_at = datetime.now()

@dataclass
class ComplianceViolation:
    """Нарушение соответствия"""
    id: str
    rule_id: str
    severity: ViolationSeverity
    description: str
    details: Dict[str, Any]
    detected_at: datetime
    resolved_at: Optional[datetime] = None
    resolution_notes: Optional[str] = None
    is_resolved: bool = False

@dataclass
class RegulatoryReport:
    """Регуляторный отчет"""
    id: str
    type: ReportType
    regulator: RegulatorType
    period_start: datetime
    period_end: datetime
    data: Dict[str, Any]
    generated_at: datetime
    file_path: Optional[str] = None
    submitted_at: Optional[datetime] = None
    submission_reference: Optional[str] = None

@dataclass
class AuditTrail:
    """Аудиторский след"""
    id: str
    user_id: str
    action: str
    entity_type: str
    entity_id: str
    old_value: Optional[Dict[str, Any]]
    new_value: Optional[Dict[str, Any]]
    timestamp: datetime
    ip_address: Optional[str] = None
    user_agent: Optional[str] = None

@dataclass
class KYCRecord:
    """Запись KYC"""
    id: str
    user_id: str
    document_type: str
    document_number: str
    verification_status: str
    verification_date: datetime
    expiry_date: Optional[datetime]
    risk_score: float
    notes: Optional[str] = None

# Метрики Prometheus
COMPLIANCE_VIOLATIONS = Counter('compliance_violations_total', 'Total compliance violations', ['rule_type', 'severity'])
REPORTS_GENERATED = Counter('regulatory_reports_total', 'Total regulatory reports generated', ['report_type', 'regulator'])
AUDIT_EVENTS = Counter('audit_events_total', 'Total audit events', ['action', 'entity_type'])
KYC_VERIFICATIONS = Counter('kyc_verifications_total', 'Total KYC verifications', ['status'])

class EncryptionManager:
    """Менеджер шифрования для конфиденциальных данных"""
    
    def __init__(self, password: str):
        self.password = password.encode()
        self.salt = b'compliance_salt_2024'  # В продакшене должен быть случайным
        self.key = self._derive_key()
        self.cipher = Fernet(self.key)
        
    def _derive_key(self) -> bytes:
        """Получение ключа шифрования"""
        kdf = PBKDF2HMAC(
            algorithm=hashes.SHA256(),
            length=32,
            salt=self.salt,
            iterations=100000,
        )
        key = base64.urlsafe_b64encode(kdf.derive(self.password))
        return key
        
    def encrypt(self, data: str) -> str:
        """Шифрование данных"""
        encrypted_data = self.cipher.encrypt(data.encode())
        return base64.urlsafe_b64encode(encrypted_data).decode()
        
    def decrypt(self, encrypted_data: str) -> str:
        """Расшифровка данных"""
        encrypted_bytes = base64.urlsafe_b64decode(encrypted_data.encode())
        decrypted_data = self.cipher.decrypt(encrypted_bytes)
        return decrypted_data.decode()

class ComplianceRuleEngine:
    """Движок правил соответствия"""
    
    def __init__(self):
        self.rules: Dict[str, ComplianceRule] = {}
        self.violations: List[ComplianceViolation] = []
        
    def add_rule(self, rule: ComplianceRule):
        """Добавление правила"""
        self.rules[rule.id] = rule
        
    def check_position_limits(self, positions: List[Dict[str, Any]]) -> List[ComplianceViolation]:
        """Проверка лимитов позиций"""
        violations = []
        
        for rule in self.rules.values():
            if rule.type == ComplianceRuleType.POSITION_LIMIT and rule.is_active:
                for position in positions:
                    symbol = position.get('symbol')
                    size = abs(position.get('size', 0))
                    
                    # Проверка лимита по символу
                    if symbol in rule.parameters.get('symbol_limits', {}):
                        limit = rule.parameters['symbol_limits'][symbol]
                        if size > limit:
                            violation = ComplianceViolation(
                                id=str(uuid.uuid4()),
                                rule_id=rule.id,
                                severity=ViolationSeverity.HIGH,
                                description=f"Position limit exceeded for {symbol}",
                                details={
                                    'symbol': symbol,
                                    'current_size': size,
                                    'limit': limit,
                                    'excess': size - limit
                                },
                                detected_at=datetime.now()
                            )
                            violations.append(violation)
                            
        return violations
        
    def check_risk_limits(self, portfolio_data: Dict[str, Any]) -> List[ComplianceViolation]:
        """Проверка лимитов риска"""
        violations = []
        
        for rule in self.rules.values():
            if rule.type == ComplianceRuleType.RISK_LIMIT and rule.is_active:
                # Проверка VaR
                if 'var_limit' in rule.parameters:
                    current_var = abs(portfolio_data.get('var_1d', 0))
                    var_limit = rule.parameters['var_limit']
                    
                    if current_var > var_limit:
                        violation = ComplianceViolation(
                            id=str(uuid.uuid4()),
                            rule_id=rule.id,
                            severity=ViolationSeverity.HIGH,
                            description="VaR limit exceeded",
                            details={
                                'current_var': current_var,
                                'limit': var_limit,
                                'excess': current_var - var_limit
                            },
                            detected_at=datetime.now()
                        )
                        violations.append(violation)
                        
                # Проверка максимальной просадки
                if 'drawdown_limit' in rule.parameters:
                    current_drawdown = abs(portfolio_data.get('max_drawdown', 0))
                    drawdown_limit = rule.parameters['drawdown_limit']
                    
                    if current_drawdown > drawdown_limit:
                        violation = ComplianceViolation(
                            id=str(uuid.uuid4()),
                            rule_id=rule.id,
                            severity=ViolationSeverity.MEDIUM,
                            description="Maximum drawdown limit exceeded",
                            details={
                                'current_drawdown': current_drawdown,
                                'limit': drawdown_limit,
                                'excess': current_drawdown - drawdown_limit
                            },
                            detected_at=datetime.now()
                        )
                        violations.append(violation)
                        
        return violations
        
    def check_trading_patterns(self, trades: List[Dict[str, Any]]) -> List[ComplianceViolation]:
        """Проверка торговых паттернов на предмет злоупотреблений"""
        violations = []
        
        # Проверка на wash trading (фиктивные сделки)
        wash_trades = self._detect_wash_trading(trades)
        if wash_trades:
            violation = ComplianceViolation(
                id=str(uuid.uuid4()),
                rule_id="market_abuse_001",
                severity=ViolationSeverity.CRITICAL,
                description="Potential wash trading detected",
                details={
                    'suspicious_trades': len(wash_trades),
                    'trade_ids': [t['id'] for t in wash_trades[:10]]  # Первые 10
                },
                detected_at=datetime.now()
            )
            violations.append(violation)
            
        # Проверка на spoofing (ложные заявки)
        spoofing_patterns = self._detect_spoofing(trades)
        if spoofing_patterns:
            violation = ComplianceViolation(
                id=str(uuid.uuid4()),
                rule_id="market_abuse_002",
                severity=ViolationSeverity.HIGH,
                description="Potential spoofing pattern detected",
                details={
                    'patterns_count': len(spoofing_patterns),
                    'details': spoofing_patterns[:5]  # Первые 5 паттернов
                },
                detected_at=datetime.now()
            )
            violations.append(violation)
            
        return violations
        
    def _detect_wash_trading(self, trades: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """Обнаружение фиктивных сделок"""
        suspicious_trades = []
        
        # Группировка сделок по времени и символу
        trade_groups = defaultdict(list)
        for trade in trades:
            key = (trade.get('symbol'), trade.get('timestamp', 0) // 60)  # Группировка по минутам
            trade_groups[key].append(trade)
            
        # Поиск подозрительных паттернов
        for group in trade_groups.values():
            if len(group) >= 2:
                # Проверка на одинаковые объемы и противоположные направления
                buy_trades = [t for t in group if t.get('side') == 'buy']
                sell_trades = [t for t in group if t.get('side') == 'sell']
                
                for buy_trade in buy_trades:
                    for sell_trade in sell_trades:
                        if (abs(buy_trade.get('quantity', 0) - sell_trade.get('quantity', 0)) < 0.001 and
                            abs(buy_trade.get('price', 0) - sell_trade.get('price', 0)) < 0.01):
                            suspicious_trades.extend([buy_trade, sell_trade])
                            
        return suspicious_trades
        
    def _detect_spoofing(self, trades: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """Обнаружение спуфинга"""
        patterns = []
        
        # Анализ отмененных заявок
        cancelled_orders = [t for t in trades if t.get('status') == 'cancelled']
        executed_orders = [t for t in trades if t.get('status') == 'filled']
        
        if len(cancelled_orders) > len(executed_orders) * 5:  # Более 5:1 отношение
            patterns.append({
                'type': 'high_cancellation_ratio',
                'cancelled_count': len(cancelled_orders),
                'executed_count': len(executed_orders),
                'ratio': len(cancelled_orders) / max(len(executed_orders), 1)
            })
            
        return patterns

class ReportGenerator:
    """Генератор регуляторных отчетов"""
    
    def __init__(self, output_dir: str = "reports"):
        self.output_dir = output_dir
        Path(output_dir).mkdir(exist_ok=True)
        
    def generate_daily_trading_report(self, trades: List[Dict[str, Any]], date: datetime) -> RegulatoryReport:
        """Генерация ежедневного торгового отчета"""
        report_id = f"daily_trading_{date.strftime('%Y%m%d')}"
        
        # Анализ торговых данных
        total_trades = len(trades)
        total_volume = sum([t.get('quantity', 0) * t.get('price', 0) for t in trades])
        symbols_traded = len(set([t.get('symbol') for t in trades]))
        
        # Группировка по символам
        symbol_stats = defaultdict(lambda: {'count': 0, 'volume': 0, 'pnl': 0})
        for trade in trades:
            symbol = trade.get('symbol')
            symbol_stats[symbol]['count'] += 1
            symbol_stats[symbol]['volume'] += trade.get('quantity', 0) * trade.get('price', 0)
            symbol_stats[symbol]['pnl'] += trade.get('pnl', 0)
            
        report_data = {
            'date': date.strftime('%Y-%m-%d'),
            'summary': {
                'total_trades': total_trades,
                'total_volume': total_volume,
                'symbols_traded': symbols_traded,
                'net_pnl': sum([t.get('pnl', 0) for t in trades])
            },
            'symbol_breakdown': dict(symbol_stats),
            'trades': trades
        }
        
        # Создание PDF отчета
        file_path = self._create_pdf_report(report_data, report_id, "Daily Trading Report")
        
        report = RegulatoryReport(
            id=report_id,
            type=ReportType.DAILY_TRADING,
            regulator=RegulatorType.SEC,  # Пример
            period_start=date,
            period_end=date + timedelta(days=1),
            data=report_data,
            generated_at=datetime.now(),
            file_path=file_path
        )
        
        return report
        
    def generate_risk_report(self, portfolio_data: Dict[str, Any], period_start: datetime, period_end: datetime) -> RegulatoryReport:
        """Генерация отчета о рисках"""
        report_id = f"risk_report_{period_start.strftime('%Y%m%d')}_{period_end.strftime('%Y%m%d')}"
        
        report_data = {
            'period': {
                'start': period_start.strftime('%Y-%m-%d'),
                'end': period_end.strftime('%Y-%m-%d')
            },
            'risk_metrics': {
                'var_1d': portfolio_data.get('var_1d', 0),
                'var_1w': portfolio_data.get('var_1w', 0),
                'max_drawdown': portfolio_data.get('max_drawdown', 0),
                'sharpe_ratio': portfolio_data.get('sharpe_ratio', 0),
                'beta': portfolio_data.get('beta', 0),
                'correlation_market': portfolio_data.get('correlation_market', 0)
            },
            'position_analysis': portfolio_data.get('positions', []),
            'stress_test_results': portfolio_data.get('stress_tests', {})
        }
        
        # Создание Excel отчета
        file_path = self._create_excel_report(report_data, report_id, "Risk Assessment Report")
        
        report = RegulatoryReport(
            id=report_id,
            type=ReportType.WEEKLY_RISK,
            regulator=RegulatorType.CFTC,
            period_start=period_start,
            period_end=period_end,
            data=report_data,
            generated_at=datetime.now(),
            file_path=file_path
        )
        
        return report
        
    def _create_pdf_report(self, data: Dict[str, Any], report_id: str, title: str) -> str:
        """Создание PDF отчета"""
        file_path = f"{self.output_dir}/{report_id}.pdf"
        
        doc = SimpleDocTemplate(file_path, pagesize=A4)
        styles = getSampleStyleSheet()
        story = []
        
        # Заголовок
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            alignment=1  # Центрирование
        )
        story.append(Paragraph(title, title_style))
        story.append(Spacer(1, 12))
        
        # Информация о периоде
        if 'date' in data:
            story.append(Paragraph(f"<b>Date:</b> {data['date']}", styles['Normal']))
        elif 'period' in data:
            story.append(Paragraph(f"<b>Period:</b> {data['period']['start']} to {data['period']['end']}", styles['Normal']))
            
        story.append(Spacer(1, 12))
        
        # Сводная информация
        if 'summary' in data:
            story.append(Paragraph("<b>Summary</b>", styles['Heading2']))
            summary_data = []
            for key, value in data['summary'].items():
                summary_data.append([key.replace('_', ' ').title(), str(value)])
                
            summary_table = Table(summary_data)
            summary_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 14),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            story.append(summary_table)
            story.append(Spacer(1, 12))
            
        # Построение документа
        doc.build(story)
        return file_path
        
    def _create_excel_report(self, data: Dict[str, Any], report_id: str, title: str) -> str:
        """Создание Excel отчета"""
        file_path = f"{self.output_dir}/{report_id}.xlsx"
        
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Report"
        
        # Стили
        header_font = Font(bold=True, size=14)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Заголовок
        worksheet['A1'] = title
        worksheet['A1'].font = Font(bold=True, size=16)
        worksheet.merge_cells('A1:D1')
        
        row = 3
        
        # Период
        if 'period' in data:
            worksheet[f'A{row}'] = "Period:"
            worksheet[f'B{row}'] = f"{data['period']['start']} to {data['period']['end']}"
            row += 2
            
        # Метрики риска
        if 'risk_metrics' in data:
            worksheet[f'A{row}'] = "Risk Metrics"
            worksheet[f'A{row}'].font = header_font
            row += 1
            
            for metric, value in data['risk_metrics'].items():
                worksheet[f'A{row}'] = metric.replace('_', ' ').title()
                worksheet[f'B{row}'] = value
                row += 1
                
        workbook.save(file_path)
        return file_path

class AuditManager:
    """Менеджер аудиторских записей"""
    
    def __init__(self, encryption_manager: EncryptionManager):
        self.encryption_manager = encryption_manager
        self.audit_trail: List[AuditTrail] = []
        
    def log_action(self, user_id: str, action: str, entity_type: str, entity_id: str,
                   old_value: Optional[Dict[str, Any]] = None, new_value: Optional[Dict[str, Any]] = None,
                   ip_address: Optional[str] = None, user_agent: Optional[str] = None):
        """Логирование действия"""
        audit_record = AuditTrail(
            id=str(uuid.uuid4()),
            user_id=user_id,
            action=action,
            entity_type=entity_type,
            entity_id=entity_id,
            old_value=old_value,
            new_value=new_value,
            timestamp=datetime.now(),
            ip_address=ip_address,
            user_agent=user_agent
        )
        
        self.audit_trail.append(audit_record)
        AUDIT_EVENTS.labels(action=action, entity_type=entity_type).inc()
        
    def get_audit_trail(self, entity_type: Optional[str] = None, entity_id: Optional[str] = None,
                       user_id: Optional[str] = None, start_date: Optional[datetime] = None,
                       end_date: Optional[datetime] = None) -> List[AuditTrail]:
        """Получение аудиторского следа с фильтрацией"""
        filtered_trail = self.audit_trail
        
        if entity_type:
            filtered_trail = [r for r in filtered_trail if r.entity_type == entity_type]
            
        if entity_id:
            filtered_trail = [r for r in filtered_trail if r.entity_id == entity_id]
            
        if user_id:
            filtered_trail = [r for r in filtered_trail if r.user_id == user_id]
            
        if start_date:
            filtered_trail = [r for r in filtered_trail if r.timestamp >= start_date]
            
        if end_date:
            filtered_trail = [r for r in filtered_trail if r.timestamp <= end_date]
            
        return filtered_trail

class KYCManager:
    """Менеджер KYC (Know Your Customer)"""
    
    def __init__(self, encryption_manager: EncryptionManager):
        self.encryption_manager = encryption_manager
        self.kyc_records: Dict[str, KYCRecord] = {}
        
    def create_kyc_record(self, user_id: str, document_type: str, document_number: str,
                         verification_status: str = "pending", expiry_date: Optional[datetime] = None,
                         notes: Optional[str] = None) -> KYCRecord:
        """Создание записи KYC"""
        # Шифрование конфиденциальных данных
        encrypted_document_number = self.encryption_manager.encrypt(document_number)
        
        # Расчет риск-скора
        risk_score = self._calculate_risk_score(user_id, document_type)
        
        kyc_record = KYCRecord(
            id=str(uuid.uuid4()),
            user_id=user_id,
            document_type=document_type,
            document_number=encrypted_document_number,
            verification_status=verification_status,
            verification_date=datetime.now(),
            expiry_date=expiry_date,
            risk_score=risk_score,
            notes=notes
        )
        
        self.kyc_records[kyc_record.id] = kyc_record
        KYC_VERIFICATIONS.labels(status=verification_status).inc()
        
        return kyc_record
        
    def update_verification_status(self, kyc_id: str, status: str, notes: Optional[str] = None):
        """Обновление статуса верификации"""
        if kyc_id in self.kyc_records:
            old_status = self.kyc_records[kyc_id].verification_status
            self.kyc_records[kyc_id].verification_status = status
            self.kyc_records[kyc_id].verification_date = datetime.now()
            
            if notes:
                self.kyc_records[kyc_id].notes = notes
                
            KYC_VERIFICATIONS.labels(status=status).inc()
            
    def _calculate_risk_score(self, user_id: str, document_type: str) -> float:
        """Расчет риск-скора пользователя"""
        base_score = 5.0  # Базовый скор
        
        # Факторы риска
        if document_type in ['passport', 'national_id']:
            base_score -= 1.0  # Надежные документы
        elif document_type in ['utility_bill', 'bank_statement']:
            base_score += 0.5  # Менее надежные документы
            
        # Дополнительные проверки (геолокация, история и т.д.)
        # В реальной системе здесь были бы более сложные алгоритмы
        
        return max(0.0, min(10.0, base_score))  # Ограничение от 0 до 10

class EnterpriseRegulatoryComplianceSystem:
    """Enterprise система регуляторного соответствия"""
    
    def __init__(self, config: Dict[str, Any]):
        self.config = config
        self.redis_client = None
        self.logger = self._setup_logging()
        
        # Компоненты системы
        self.encryption_manager = EncryptionManager(config.get('encryption_password', 'default_password'))
        self.rule_engine = ComplianceRuleEngine()
        self.report_generator = ReportGenerator(config.get('reports_dir', 'reports'))
        self.audit_manager = AuditManager(self.encryption_manager)
        self.kyc_manager = KYCManager(self.encryption_manager)
        
        # Состояние системы
        self.active_violations: List[ComplianceViolation] = []
        self.generated_reports: List[RegulatoryReport] = []
        
        # Инициализация правил по умолчанию
        self._initialize_default_rules()
        
    def _setup_logging(self) -> logging.Logger:
        """Настройка логирования"""
        logger = logging.getLogger('enterprise_compliance')
        logger.setLevel(logging.INFO)
        
        handler = logging.StreamHandler()
        formatter = logging.Formatter(
            '%(asctime)s - %(name)s - %(levelname)s - %(message)s'
        )
        handler.setFormatter(formatter)
        logger.addHandler(handler)
        
        return logger
        
    def _initialize_default_rules(self):
        """Инициализация правил по умолчанию"""
        # Правило лимитов позиций
        position_rule = ComplianceRule(
            id="position_limit_001",
            name="Position Size Limits",
            type=ComplianceRuleType.POSITION_LIMIT,
            regulator=RegulatorType.SEC,
            description="Maximum position sizes per symbol",
            parameters={
                'symbol_limits': {
                    'BTCUSDT': 10.0,
                    'ETHUSDT': 100.0,
                    'ADAUSDT': 10000.0
                }
            }
        )
        self.rule_engine.add_rule(position_rule)
        
        # Правило лимитов риска
        risk_rule = ComplianceRule(
            id="risk_limit_001",
            name="Risk Limits",
            type=ComplianceRuleType.RISK_LIMIT,
            regulator=RegulatorType.CFTC,
            description="Maximum risk exposure limits",
            parameters={
                'var_limit': 5000.0,  # Максимальный VaR
                'drawdown_limit': 5.0  # Максимальная просадка в %
            }
        )
        self.rule_engine.add_rule(risk_rule)
        
    async def start(self):
        """Запуск системы соответствия"""
        # Подключение к Redis
        self.redis_client = redis.Redis(
            host=self.config.get('redis_host', 'localhost'),
            port=self.config.get('redis_port', 6379),
            decode_responses=True
        )
        
        # Запуск фоновых задач
        asyncio.create_task(self._continuous_monitoring())
        asyncio.create_task(self._scheduled_reporting())
        asyncio.create_task(self._violation_alerts())
        
        self.logger.info("Enterprise Regulatory Compliance System started")
        
    async def stop(self):
        """Остановка системы"""
        if self.redis_client:
            await self.redis_client.close()
            
    async def run_compliance_check(self, trading_data: Dict[str, Any]) -> Dict[str, Any]:
        """Запуск проверки соответствия"""
        self.logger.info("Running compliance check")
        
        violations = []
        
        try:
            # Проверка лимитов позиций
            if 'positions' in trading_data:
                position_violations = self.rule_engine.check_position_limits(trading_data['positions'])
                violations.extend(position_violations)
                
            # Проверка лимитов риска
            if 'portfolio' in trading_data:
                risk_violations = self.rule_engine.check_risk_limits(trading_data['portfolio'])
                violations.extend(risk_violations)
                
            # Проверка торговых паттернов
            if 'trades' in trading_data:
                pattern_violations = self.rule_engine.check_trading_patterns(trading_data['trades'])
                violations.extend(pattern_violations)
                
            # Сохранение нарушений
            for violation in violations:
                self.active_violations.append(violation)
                COMPLIANCE_VIOLATIONS.labels(
                    rule_type=violation.rule_id.split('_')[0],
                    severity=violation.severity.value
                ).inc()
                
                # Логирование в аудиторский след
                self.audit_manager.log_action(
                    user_id="system",
                    action="compliance_violation_detected",
                    entity_type="violation",
                    entity_id=violation.id,
                    new_value=asdict(violation)
                )
                
            # Сохранение в Redis
            await self.redis_client.set(
                "compliance_check_result",
                json.dumps({
                    'timestamp': datetime.now().isoformat(),
                    'violations_count': len(violations),
                    'violations': [asdict(v) for v in violations]
                }),
                ex=3600
            )
            
            result = {
                'status': 'completed',
                'violations_found': len(violations),
                'violations': violations,
                'timestamp': datetime.now().isoformat()
            }
            
            self.logger.info(f"Compliance check completed. Found {len(violations)} violations")
            return result
            
        except Exception as e:
            self.logger.error(f"Compliance check error: {e}")
            return {
                'status': 'error',
                'error': str(e),
                'timestamp': datetime.now().isoformat()
            }
            
    async def generate_regulatory_report(self, report_type: ReportType, period_start: datetime,
                                       period_end: datetime, data: Dict[str, Any]) -> RegulatoryReport:
        """Генерация регуляторного отчета"""
        self.logger.info(f"Generating {report_type.value} report for period {period_start} to {period_end}")
        
        try:
            if report_type == ReportType.DAILY_TRADING:
                report = self.report_generator.generate_daily_trading_report(
                    data.get('trades', []), period_start
                )
            elif report_type == ReportType.WEEKLY_RISK:
                report = self.report_generator.generate_risk_report(
                    data.get('portfolio', {}), period_start, period_end
                )
            else:
                raise ValueError(f"Unsupported report type: {report_type}")
                
            # Сохранение отчета
            self.generated_reports.append(report)
            REPORTS_GENERATED.labels(
                report_type=report_type.value,
                regulator=report.regulator.value
            ).inc()
            
            # Логирование
            self.audit_manager.log_action(
                user_id="system",
                action="report_generated",
                entity_type="report",
                entity_id=report.id,
                new_value=asdict(report)
            )
            
            # Сохранение в Redis
            await self.redis_client.set(
                f"report_{report.id}",
                json.dumps(asdict(report), default=str),
                ex=86400 * 30  # 30 дней
            )
            
            self.logger.info(f"Report {report.id} generated successfully")
            return report
            
        except Exception as e:
            self.logger.error(f"Report generation error: {e}")
            raise
            
    async def submit_report_to_regulator(self, report_id: str, submission_method: str = "email") -> bool:
        """Отправка отчета регулятору"""
        report = next((r for r in self.generated_reports if r.id == report_id), None)
        if not report:
            raise ValueError(f"Report {report_id} not found")
            
        try:
            if submission_method == "email":
                success = await self._submit_report_by_email(report)
            else:
                raise ValueError(f"Unsupported submission method: {submission_method}")
                
            if success:
                report.submitted_at = datetime.now()
                report.submission_reference = f"REF_{int(time.time())}"
                
                # Логирование
                self.audit_manager.log_action(
                    user_id="system",
                    action="report_submitted",
                    entity_type="report",
                    entity_id=report.id,
                    new_value={'submitted_at': report.submitted_at.isoformat()}
                )
                
            return success
            
        except Exception as e:
            self.logger.error(f"Report submission error: {e}")
            return False
            
    async def _submit_report_by_email(self, report: RegulatoryReport) -> bool:
        """Отправка отчета по email"""
        try:
            # Настройки email (в реальной системе из конфигурации)
            smtp_server = self.config.get('smtp_server', 'smtp.gmail.com')
            smtp_port = self.config.get('smtp_port', 587)
            email_user = self.config.get('email_user', 'compliance@company.com')
            email_password = self.config.get('email_password', 'password')
            
            # Адреса регуляторов
            regulator_emails = {
                RegulatorType.SEC: 'reports@sec.gov',
                RegulatorType.CFTC: 'reports@cftc.gov',
                RegulatorType.FINRA: 'reports@finra.org'
            }
            
            recipient = regulator_emails.get(report.regulator)
            if not recipient:
                self.logger.warning(f"No email configured for regulator {report.regulator}")
                return False
                
            # Создание сообщения
            msg = MIMEMultipart()
            msg['From'] = email_user
            msg['To'] = recipient
            msg['Subject'] = f"Regulatory Report - {report.type.value} - {report.id}"
            
            body = f"""
Dear Regulatory Authority,

Please find attached the {report.type.value} report for the period from {report.period_start} to {report.period_end}.

Report ID: {report.id}
Generated: {report.generated_at}
Report Type: {report.type.value}

Best regards,
Compliance Department
            """
            
            msg.attach(MIMEText(body, 'plain'))
            
            # Прикрепление файла отчета
            if report.file_path and Path(report.file_path).exists():
                with open(report.file_path, "rb") as attachment:
                    part = MIMEBase('application', 'octet-stream')
                    part.set_payload(attachment.read())
                    
                encoders.encode_base64(part)
                part.add_header(
                    'Content-Disposition',
                    f'attachment; filename= {Path(report.file_path).name}'
                )
                msg.attach(part)
                
            # Отправка
            server = smtplib.SMTP(smtp_server, smtp_port)
            server.starttls()
            server.login(email_user, email_password)
            text = msg.as_string()
            server.sendmail(email_user, recipient, text)
            server.quit()
            
            self.logger.info(f"Report {report.id} sent to {recipient}")
            return True
            
        except Exception as e:
            self.logger.error(f"Email submission error: {e}")
            return False
            
    async def _continuous_monitoring(self):
        """Непрерывный мониторинг соответствия"""
        while True:
            try:
                # Получение текущих торговых данных
                trading_data_json = await self.redis_client.get("current_trading_data")
                if trading_data_json:
                    trading_data = json.loads(trading_data_json)
                    await self.run_compliance_check(trading_data)
                    
                await asyncio.sleep(60)  # Проверка каждую минуту
                
            except Exception as e:
                self.logger.error(f"Continuous monitoring error: {e}")
                await asyncio.sleep(300)  # Пауза при ошибке
                
    async def _scheduled_reporting(self):
        """Запланированная отчетность"""
        while True:
            try:
                now = datetime.now()
                
                # Ежедневные отчеты в 23:00
                if now.hour == 23 and now.minute == 0:
                    trading_data_json = await self.redis_client.get("daily_trading_data")
                    if trading_data_json:
                        trading_data = json.loads(trading_data_json)
                        await self.generate_regulatory_report(
                            ReportType.DAILY_TRADING,
                            now.replace(hour=0, minute=0, second=0, microsecond=0),
                            now,
                            trading_data
                        )
                        
                # Еженедельные отчеты в воскресенье в 22:00
                if now.weekday() == 6 and now.hour == 22 and now.minute == 0:
                    portfolio_data_json = await self.redis_client.get("portfolio_data")
                    if portfolio_data_json:
                        portfolio_data = json.loads(portfolio_data_json)
                        week_start = now - timedelta(days=7)
                        await self.generate_regulatory_report(
                            ReportType.WEEKLY_RISK,
                            week_start,
                            now,
                            {'portfolio': portfolio_data}
                        )
                        
                await asyncio.sleep(60)  # Проверка каждую минуту
                
            except Exception as e:
                self.logger.error(f"Scheduled reporting error: {e}")
                await asyncio.sleep(300)
                
    async def _violation_alerts(self):
        """Алерты о нарушениях"""
        while True:
            try:
                # Проверка критических нарушений
                critical_violations = [
                    v for v in self.active_violations 
                    if v.severity == ViolationSeverity.CRITICAL and not v.is_resolved
                ]
                
                if critical_violations:
                    # Отправка уведомлений (email, Slack и т.д.)
                    await self._send_violation_alerts(critical_violations)
                    
                await asyncio.sleep(300)  # Проверка каждые 5 минут
                
            except Exception as e:
                self.logger.error(f"Violation alerts error: {e}")
                await asyncio.sleep(600)
                
    async def _send_violation_alerts(self, violations: List[ComplianceViolation]):
        """Отправка алертов о нарушениях"""
        for violation in violations:
            self.logger.warning(f"CRITICAL VIOLATION: {violation.description}")
            # Здесь можно добавить отправку email, Slack, SMS и т.д.

async def main():
    """Основная функция запуска"""
    config = {
        'encryption_password': 'enterprise_compliance_2024',
        'reports_dir': 'compliance_reports',
        'redis_host': 'localhost',
        'redis_port': 6379,
        'smtp_server': 'smtp.gmail.com',
        'smtp_port': 587,
        'email_user': 'compliance@company.com',
        'email_password': 'app_password'
    }
    
    compliance_system = EnterpriseRegulatoryComplianceSystem(config)
    await compliance_system.start()
    
    print("Enterprise Regulatory Compliance System started")
    
    try:
        # Пример использования
        
        # Тестовые торговые данные
        test_trading_data = {
            'positions': [
                {'symbol': 'BTCUSDT', 'size': 5.0},
                {'symbol': 'ETHUSDT', 'size': 50.0}
            ],
            'portfolio': {
                'var_1d': 3000.0,
                'max_drawdown': 2.5,
                'sharpe_ratio': 2.8
            },
            'trades': [
                {'id': 'trade_1', 'symbol': 'BTCUSDT', 'side': 'buy', 'quantity': 1.0, 'price': 45000, 'pnl': 100},
                {'id': 'trade_2', 'symbol': 'BTCUSDT', 'side': 'sell', 'quantity': 1.0, 'price': 45100, 'pnl': 100}
            ]
        }
        
        # Запуск проверки соответствия
        print("Running compliance check...")
        result = await compliance_system.run_compliance_check(test_trading_data)
        print(f"Compliance check result: {result['status']}")
        print(f"Violations found: {result['violations_found']}")
        
        # Генерация отчета
        print("Generating daily trading report...")
        report = await compliance_system.generate_regulatory_report(
            ReportType.DAILY_TRADING,
            datetime.now().replace(hour=0, minute=0, second=0, microsecond=0),
            datetime.now(),
            test_trading_data
        )
        print(f"Report generated: {report.id}")
        
        # Создание KYC записи
        print("Creating KYC record...")
        kyc_record = compliance_system.kyc_manager.create_kyc_record(
            user_id="user_001",
            document_type="passport",
            document_number="AB1234567",
            verification_status="verified"
        )
        print(f"KYC record created: {kyc_record.id}")
        
        # Непрерывная работа
        print("System running... Press Ctrl+C to stop")
        await asyncio.Future()  # Бесконечное ожидание
        
    except KeyboardInterrupt:
        print("Shutting down...")
    finally:
        await compliance_system.stop()

if __name__ == '__main__':
    asyncio.run(main())